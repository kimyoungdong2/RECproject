# -*- coding: utf-8 -*-
import openpyxl
"""DATA_OF_2017={
    'First_Number' : 749, #First_Number = 우선선발 신청한 인원 수
    'Second_Number' : 448, #Second_Number = 일반선발A 신청한 인원 수
    'Third_Number' : 1, #Third_Number = 일반선발B 신청한 인원 수
    'First_Input' : 82900, #First_Input = 우선선발의 총 용량(가중치가 포함 됨)
    'Second_Input' : 403514, #Second_Input = 일반선발A의 총 용량(가중치가 포함 됨)
    'Third_Input' : 3843, #Thrid_Input = 일반선발B의 총 용량(가중치가 포함 됨)
    'First_Capacity' : 150000, #First_Capacity = 250000(공고용량) * 0.6
    'Second_Capacity' : 75000, #Second_Capacity = 250000(공고용량) * 0.3
    'Third_Capacity' : 25000, #Third_Capacity = 250000(공고용량) * 0.1
    'Upper_Limit' : 110, #2017년 SMP 평균 가격 81.125  REC 평균 가격 103,473 (육지기준)
    'Average' : 100, # 1REC = 1000kw (가중치는 이미 계산 됨) then Upper_Limit = 110,205 (SMP가 연 평균이므로 110으로 생각)
}
"""
DATA_OF_2017 = {
    'First_Number' : 1548, #First_Number = 우선선발 신청한 인원 수
    'Second_Number' : 711,  #Second_Number = 일반선발A 신청한 인원 수
    'Third_Number' : 1, #Third_Number = 일반선발B 신청한 인원 수
    'First_Input' : 171985, #First_Input = 우선선발의 총 용량(가중치가 포함 됨)
    'Second_Input' : 575856, #Second_Input = 일반선발A의 총 용량(가중치가 포함 됨)
    'Third_Input' : 3844, #Thrid_Input = 일반선발B의 총 용량(가중치가 포함 됨)
    'First_Capacity' : 150000, #First_Capacity = 250000(공고용량) * 0.6
    'Second_Capacity' : 75000, #Second_Capacity = 250000(공고용량) * 0.3
    'Third_Capacity' : 25000, #Third_Capacity = 250000(공고용량) * 0.1
    'Upper_Limit' : 110, #2017년 SMP 평균 가격 81.125  REC 평균 가격 103,473 (육지기준) then Upper_Limit = 110,205 (SMP가 연 평균이므로 110으로 생각)
    'Average' : 103, # 1REC = 1000kw (가중치는 이미 계산 됨) then  Average = 103.473 (SMP가 연 평균이므로 103으로 생각) 우선선발, 일반선발A, 일반선발B 모두 똑같다고 가정 
}
          
First_Average=float(DATA_OF_2017['First_Input'])/float(DATA_OF_2017['First_Number'])
Second_Average=float(DATA_OF_2017['Second_Input'])/float(DATA_OF_2017['Second_Number'])
Third_Average=float(DATA_OF_2017['Third_Input'])/float(DATA_OF_2017['Third_Number'])

Standard_Deviation_List =[5,10,20,30]  #표준편차  = 다양하게 계산 할 예정


wb=openpyxl.load_workbook("연습.xlsx")
ws=wb.active
ws["B1"],ws["B14"],ws["B27"],ws["B40"]="알파 퍼센트(공정성)","알파 퍼센트(공정성)","알파 퍼센트(공정성)","알파 퍼센트(공정성)"
ws["A1"],ws["A14"],ws["A27"],ws["A40"]="표준편차","표준편차","표준편차","표준편차"
ws["A2"],ws["A15"],ws["A28"],ws["A41"]="5","10","20","30"
ws["C1"],ws["C14"],ws["C27"],ws["C40"]="효율성","효율성","효율성","효율성"
ws["D1"],ws["D14"],ws["D27"],ws["D40"]="Jain Index","Jain Index","Jain Index","Jain Index"
ws["E1"],ws["E14"],ws["E27"],ws["E40"]="예정","예정","예정","예정"

for i in range(2,12):
    ws.cell(row=i,column=2).value=0.02*(i-1)
for i in range(2,12):
    ws.cell(row=i+13,column=2).value=0.02*(i-1)
for i in range(2,12):
    ws.cell(row=i+26,column=2).value=0.02*(i-1)
for i in range(2,12):
    ws.cell(row=i+39,column=2).value=0.02*(i-1)


in_advance_List=[0.02,0.04,0.06,0.08,0.10,0.12,0.14,0.16,0.18,0.20]
            
The_end=0
All_Fair=0
