# -*- coding: utf-8 -*-
#표준편차 5단위
#예정된 곳에서 팔린 REC 비율
import openpyxl
import random
import config
import module
         

for one in range(len(config.Standard_Deviation_List)):
    Standard_Deviation = config.Standard_Deviation_List[one]
    for two in range(len(config.in_advance_List)):
        in_advance=config.in_advance_List[two]
        for three in range(10):
            First_Quantitative_Volume = module.Randoms_Price(config.DATA_OF_2017['First_Number'], config.First_Average, Standard_Deviation) #정량의 파는 양(Volume) 정규분포(1)
            module.Quantitative_Volume(First_Quantitative_Volume,150,0,config.First_Average,Standard_Deviation)
            
            First_Quantitative_Price = module.Randoms_Uniform(config.DATA_OF_2017['First_Number'],config.DATA_OF_2017['Average']-(config.DATA_OF_2017['Upper_Limit']-config.DATA_OF_2017['Average']) ,config.DATA_OF_2017['Upper_Limit']) # 정량적인 가격은 앞뒤 10씩
            First_Qualitative_Price = module.Randoms_Uniform(config.DATA_OF_2017['First_Number'], 25.5, 30) #정성적인거는 25.5에서 30까지 랜덤으로
            
            Second_Quantitative_Volume = module.Randoms_Price(config.DATA_OF_2017['Second_Number'], config.Second_Average, Standard_Deviation) #정량의 파는 양(Volume) 정규분포(1)
            module.Quantitative_Volume(Second_Quantitative_Volume,4500,100,config.Second_Average,Standard_Deviation)
            
            Second_Quantitative_Price = module.Randoms_Uniform(config.DATA_OF_2017['Second_Number'], config.DATA_OF_2017['Average']-(config.DATA_OF_2017['Upper_Limit']-config.DATA_OF_2017['Average']) ,config.DATA_OF_2017['Upper_Limit']) # 정량적인 가격은 앞뒤 10씩
            Second_Qualitative_Price = module.Randoms_Uniform(config.DATA_OF_2017['Second_Number'], 25.5, 30) #정성적인거는 25.5에서 30까지 랜덤으로
            
            Third_Quantitative_Volume = module.Randoms_Price(config.DATA_OF_2017['Third_Number'], config.Third_Average, Standard_Deviation) #정량의 파는 양(Volume) 정규분포(1)
            module.Quantitative_Volume(Third_Quantitative_Volume,2100,1000000,config.Third_Average,Standard_Deviation)   #1000000은 굉장히 큰 수
                        
            Third_Quantitative_Price = module.Randoms_Uniform(config.DATA_OF_2017['Third_Number'], config.DATA_OF_2017['Average']-(config.DATA_OF_2017['Upper_Limit']-config.DATA_OF_2017['Average']) ,config.DATA_OF_2017['Upper_Limit']) # 정량적인 가격은 앞뒤 10씩
            Third_Qualitative_Price = module.Randoms_Uniform(config.DATA_OF_2017['Third_Number'], 25.5, 30) #정성적인거는 25.5에서 30까지 랜덤으로
            
            
            First_All = module.All(config.DATA_OF_2017['First_Number'],First_Quantitative_Price,First_Quantitative_Volume,First_Qualitative_Price)
            
            Second_All = module.All(config.DATA_OF_2017['Second_Number'],Second_Quantitative_Price,Second_Quantitative_Volume,Second_Qualitative_Price)
            
            Third_All = module.All(config.DATA_OF_2017['Third_Number'],Third_Quantitative_Price,Third_Quantitative_Volume,Third_Qualitative_Price)
            
            #효율성의 efficiency은 우선선발, 일반선발A, 일반선발B를 다 통합해서 그것들 중에 가격이 낮은거(Quantitative_Price가 큰거)부터 구매해서 효율성 최대화
            Down_for_efficiency=0 #효율성 계산을 위한 분모
            All_of_first_second_third=[] #First, Second, Third에 있는 모든것들을 다 합한 것
            All_of_first_second_third=First_All+Second_All+Third_All
            All_of_first_second_third.sort(reverse=True,key=lambda x:x[0])
            tmp_for_efficiency=0 #바로 밑에 for문에서만 사용하고 의미 없는 것
            for i in range(len(All_of_first_second_third)):  #Down_for_efficiency에 가격이 낮은거(Quantitative_Price가 큰거) * 그에 맞는 양(Volume)을 최대한 넣는 작업
                if tmp_for_efficiency< config.DATA_OF_2017['First_Capacity']+config.DATA_OF_2017['Second_Capacity']+config.DATA_OF_2017['Third_Capacity']: #총 판매 용량을 넘으면 안 됨.
                    Down_for_efficiency+=All_of_first_second_third[i][0]*All_of_first_second_third[i][1]
                    tmp_for_efficiency+=All_of_first_second_third[i][0]
                else:
                    break
            
            
            First_Middle_Sum=in_advance*sum((First_All[i][1]) for i in range(len(First_All)))   # First_Middle_Sum = 몇 %를 먼저 구할것인지.
            Second_Middle_Sum=in_advance*sum((Second_All[i][1]) for i in range(len(Second_All)))  # Second_Middle_Sum = 몇 %를 먼저 구할것인지.
            Third_Middle_Sum=in_advance*sum((Third_All[i][1]) for i in range(len(Third_All)))   # Third_Middle_Sum = 몇 %를 먼저 구할것인지.
                
            
            module.a_persent(First_All,in_advance)
            module.a_persent(Second_All,in_advance)
            module.a_persent(Third_All,in_advance)
            

            module.Sum(config.DATA_OF_2017['First_Number'],First_Middle_Sum,First_All)
            
            First_Middle_Pass = First_All[:i+1]  #1.3배안에 속한 판매자
            First_Middle_Fail = First_All[i+1:]  #1.3배안에 속하지 못한 판매자
            
            module.Change(First_Middle_Pass,config.DATA_OF_2017['Upper_Limit'])            
            
            First_Sum = module.Selected(First_Middle_Pass,config.DATA_OF_2017['First_Capacity'],First_Middle_Sum)
            
            First_Final_Pass = First_Middle_Pass[:i+1] #우선선정에 최종 선발자
            First_Final_Fail = First_Middle_Pass[i+1:] #우선선정 마지막 단계 탈락자
            
            First_fail_plus_second = Second_All + First_Middle_Fail + First_Final_Fail   #우선선발에서 실패한 사람들이 모두 내려옴
            First_fail_plus_second.sort(reverse=True)  #가격이 낮은 순서대로 정렬


            module.Sum(config.DATA_OF_2017['Second_Number'],Second_Middle_Sum,Second_All)
            
            Second_Middle_Pass = First_fail_plus_second[:i+1] #1.3배안에 속한 판매자
            Second_Middle_Fail = First_fail_plus_second[i+1:] #1.3배안에 속하지 못한 판매자
            
            module.Change(Second_Middle_Pass,config.DATA_OF_2017['Upper_Limit'])            
            
            Second_Sum = module.Selected(Second_Middle_Pass,config.DATA_OF_2017['Second_Capacity'],Second_Middle_Sum)

            Second_Final_Pass = Second_Middle_Pass[:i+1] #일반선발A에 최종 선발자
            Second_Final_Fail = Second_Middle_Pass[i+1:] #일반선발A에 마지막 단계 탈락자
            
            First_fail_plus_second_fail_plus_third = Third_All + Second_Middle_Fail + Second_Final_Fail #일반선발A에서 실패한사람들이 모두 내려옴
            First_fail_plus_second_fail_plus_third.sort(reverse=True)   #가격이 낮은 순서대로 정렬
            

            module.Sum(config.DATA_OF_2017['Third_Number'],Third_Middle_Sum,Third_All)
            
            Third_Middle_Pass = First_fail_plus_second_fail_plus_third[:i+1] #1.3배안에 속한 판매자
            Third_Middle_Fail = First_fail_plus_second_fail_plus_third[i+1:] #1.3배안에 속하지 못한 판매자
            
            module.Change(Third_Middle_Pass,config.DATA_OF_2017['Upper_Limit'])            
            
            Third_Sum = module.Selected(Third_Middle_Pass,config.DATA_OF_2017['Third_Capacity'],Third_Middle_Sum)
            
            Third_Final_Pass = Third_Middle_Pass[:i] #일반선발B에 최종 선발자
            Third_Final_Fail = Third_Middle_Pass[i:] #일반선발B에 마지막 단계 탈락자
            
            First_All_Fail = sorted(First_Middle_Fail[:] + First_Final_Fail[:], reverse=True)  #우선선정에서 탈락자들
            Second_All_Fail = sorted(Second_Middle_Fail[:] + Second_Final_Fail[:], reverse=True)  #일반선정A에서 탈락자들
            Third_All_Fail = sorted(Third_Middle_Fail[:] + Third_Final_Fail[:], reverse=True)   #일반선정B에서 탈락자들
            
            
            Final_Pass = sorted(First_Final_Pass + Second_Final_Pass + Third_Final_Pass, reverse=True)   #다 합치기
            First_Get = []    #우선선정에서 뽑힌 사람들이 진짜 우선선정에서 뽑혔는지 알기 위해 변수 생성
            Second_Get = []    #일반선정A에서 뽑힌 사람들이 진짜 일반선정A에서 뽑혔는지 알기 위해 변수 생성
            Third_Get = []    #일반선정B에서 뽑힌 사람들이 진짜 일반선정B에서 뽑혔는지 알기 위해 변수 생성
            
            for i in range(len(Final_Pass)):      #다 합친거에서 출신이 어디냐로 다시 나눔
                if Final_Pass[i] in First_All:
                    First_Get.append(Final_Pass[i])
                elif Final_Pass[i] in Second_All:
                    Second_Get.append(Final_Pass[i])
                else:
                    Third_Get.append(Final_Pass[i])
            
                    
            First_jain_xi=module.Jain(First_Get,First_All,in_advance)                   # 실제 측정된(y)/이상적인(z)의 비율 즉, 뽑힌거는 1 안뽑힌거는 in_advance를 넣음   in_advance만큼은 무조건 팔림                                              
                
            First_jain_Up=sum(First_jain_xi)**2   
            First_jain_Down=len(First_jain_xi)*sum(i**2 for i in First_jain_xi)
            First_Fairness_Index=0 if First_jain_Up==0 or First_jain_Down==0 else First_jain_Up/First_jain_Down #분모나 분자가0이면 아예불공정을 의미하므로 jain_index도 0
                
            
            Second_jain_xi=module.Jain(Second_Get,Second_All,in_advance)
            
            Second_jain_Up=sum(Second_jain_xi)**2
            Second_jain_Down=len(Second_jain_xi)*sum(i**2 for i in Second_jain_xi)
            Second_Fairness_Index=0 if Second_jain_Up==0 or Second_jain_Down==0 else Second_jain_Up/Second_jain_Down #분모나 분자가0이면 아예불공정을 의미하므로 jain_index도 0
            
            Third_jain_xi=module.Jain(Third_Get,Third_All,in_advance)
            
            Third_jain_Up=sum(Third_jain_xi)**2
            Third_jain_Down=len(Third_jain_xi)*sum(i**2 for i in Third_jain_xi)
            Third_Fairness_Index=0 if Third_jain_Up==0 or Third_jain_Down==0 else Third_jain_Up/Third_jain_Down #분모나 분자가0이면 아예불공정을 의미하므로 jain_index도 0
            
            
            All_Fairness_Index=(First_Fairness_Index+Second_Fairness_Index+Third_Fairness_Index)/3 #First, Second, Third jain_index를 모두 더하고 3으로 나눔
            config.Jain_fairness_index+=All_Fairness_Index
            All_of_pass=0

            for i in range(len(First_Final_Pass)):
                All_of_pass+=First_Final_Pass[i][0]*First_Final_Pass[i][1]
            for i in range(len(Second_Final_Pass)):
                All_of_pass+=Second_Final_Pass[i][0]*Second_Final_Pass[i][1]
            for i in range(len(Third_Final_Pass)):
                All_of_pass+=Third_Final_Pass[i][0]*Third_Final_Pass[i][1]
                
            All_of_first_second_third.sort()
            The_lowest_Quantitative_price=All_of_first_second_third[0][0]   
            Up_for_efficiency = All_of_pass +(First_Middle_Sum+Second_Middle_Sum+Third_Middle_Sum)*The_lowest_Quantitative_price    #Up_for_efficiency = 총 성공량(즉, 가격*양) + (가장 낮은 가격 * 미리 선정해둔것들)
            config.efficiency+=Up_for_efficiency/Down_for_efficiency
        config.ws.cell(row=2+two+(one*13),column=3).value=config.efficiency/10
        config.ws.cell(row=2+two+(one*13),column=4).value=config.Jain_fairness_index/10
        Percentage_of_scheduled_locations=(len(First_Get)+len(Second_Get)+len(Third_Get))/(len(First_All)+len(Second_All)+len(Third_All))
        config.ws.cell(row=2+two+(one*13),column=5).value=Percentage_of_scheduled_locations
        config.efficiency=0
        config.Jain_fairness_index=0
config.wb.save("결과물.xlsx")
config.wb.close()