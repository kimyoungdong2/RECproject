import openpyxl
import random

def Randoms(x):
    total = []
    test = 0
    for i in range(x):
        test = random.random()
        total.append(test)
    return total


def Randoms_Uniform(x, Start, Finish):
    total = []
    test = 0
    for i in range(x):
        test = random.uniform(Start, Finish)
        total.append(test)
    return total

def Randoms_Gauss(Num,x,y):
    total = []
    test = 0
    for i in range(Num):
        test = random.gauss(x,y)
        if test<0:
            test=0
        elif test>Upper_Limit:
            test=Upper_Limit
        total.append(test)
    return total

def Randoms_Price(Num,x,y):
    total=[]
    test=0
    for i in range(Num):
        test=random.gauss(x,y)
        total.append(test)
    return total


First_Number = int(input("How many people want to bid in First Line?:"))
Second_Number = int(input("How many people want to bid in Second Line?:"))
Third_Number = int(input("How many people want to bid in Third Line?:"))

First_Input= int(input("How many input want to bid in First Line?:"))
Second_Input= int(input("How many input want to bid in First Line?:"))
Third_Input= int(input("How many input want to bid in First Line?:"))

First_Capacity = float(input("How many capacity does First selection Have?:"))
Second_Capacity = float(input("How many capacity does Second selection Have?:"))
Third_Capacity = float(input("How many capacity does Third selection Have?:"))

Upper_Limit = float(input("What is your upper limit?"))

First_Average=First_Input/First_Number
Second_Average=Second_Input/Second_Number
Third_Average=Third_Input/Third_Number


Average = float(input("What is the average?"))
#Standard_Deviation =float(input("What is the Standard Deviation?"))

All_Average=0
First_Middle_Index,Second_Middle_Index,Third_Middle_Index=0,0,0

wb=openpyxl.load_workbook("연습.xlsx")
ws=wb.active
ws["B1"],ws["B14"],ws["B27"],ws["B40"]="가격 표준편차","가격 표준편차","가격 표준편차","가격 표준편차"
ws["A1"],ws["A14"],ws["A27"],ws["A40"]="용량 표준편차","용량 표준편차","용량 표준편차","용량 표준편차"
ws["D1"],ws["D14"],ws["D27"],ws["D40"]="우선선발","일반선발A","일반선발B","총 평균"

for i in range(2,12):
    ws.cell(row=2,column=i).value=3*(i-1)
for i in range(2,12):
    ws.cell(row=15,column=i).value=3*(i-1)
for i in range(2,12):
    ws.cell(row=28,column=i).value=3*(i-1)
for i in range(2,12):
    ws.cell(row=41,column=i).value=3*(i-1)

K=[]
in_advance=0.05
for ss in range(10):
    Standard_Deviation=5
    First_Quantitative_Volume = Randoms_Price(First_Number, First_Average, Standard_Deviation) #양은 정규분포
    for i in range(len(First_Quantitative_Volume)): 
        if First_Quantitative_Volume[i] < 0 or First_Quantitative_Volume[i] > 150:
            First_Quantitative_Volume[i] = random.gauss(First_Average, Standard_Deviation)
    
    First_Quantitative_Price = Randoms_Uniform(First_Number,Average-(Upper_Limit-Average) ,Upper_Limit) #가격은 앞뒤 10씩
    First_Qualitative_Price = Randoms_Uniform(First_Number, 25.5, 30) #정성적인거는 랜덤으로
    
    Second_Quantitative_Volume = Randoms_Price(Second_Number, Second_Average, Standard_Deviation) #양은 정규분포
    for i in range(len(Second_Quantitative_Volume)): 
        if Second_Quantitative_Volume[i] < 100 or Second_Quantitative_Volume[i] > 4500:
            Second_Quantitative_Volume[i] = random.gauss(Second_Average, Standard_Deviation)
    
    Second_Quantitative_Price = Randoms_Uniform(Second_Number, Average-(Upper_Limit-Average) ,Upper_Limit)
    Second_Qualitative_Price = Randoms_Uniform(Second_Number, 25.5, 30)
    
    Third_Quantitative_Volume = Randoms_Price(Third_Number, Third_Average, Standard_Deviation)
    for i in range(len(Third_Quantitative_Volume)):
        if Second_Quantitative_Volume[i] < 2100:
            Second_Quantitative_Volume[i] = random.gauss(Second_Average, Standard_Deviation)
    
    Third_Quantitative_Price = Randoms_Uniform(Third_Number, Average-(Upper_Limit-Average) ,Upper_Limit)
    Third_Qualitative_Price = Randoms_Uniform(Third_Number, 25.5, 30)
    First_All = []
    for i in range(First_Number):
        First_All.append([First_Quantitative_Price[i], First_Quantitative_Volume[i], First_Qualitative_Price[i]]) #정량적인 가격을 맨 앞에 세워서 sort함
    First_All.sort(reverse=True)
    
    Second_All = []
    for i in range(Second_Number):
        Second_All.append([Second_Quantitative_Price[i], Second_Quantitative_Volume[i], Second_Qualitative_Price[i]])
    Second_All.sort(reverse=True)
    
    
    Third_All = []
    for i in range(Third_Number):
        Third_All.append([Third_Quantitative_Price[i], Third_Quantitative_Volume[i], Third_Qualitative_Price[i]])
    Third_All.sort(reverse=True)
    
    All_tmp=0
    
    
    First_Middle_Sum=in_advance*sum(float(First_All[i][1]) for i in range(len(First_All)))   #몇 %를 먼저 구할것인지.
    Second_Middle_Sum=in_advance*sum(float(Second_All[i][1]) for i in range(len(Second_All))) 
    Third_Middle_Sum=in_advance*sum(float(Third_All[i][1]) for i in range(len(Third_All)))
    Down=0
    Down_tmp=[]
    Down_tmp=First_All+Second_All+Third_All
    Down_tmp.sort(reverse=True,key=lambda x:x[0])
    for i in range(len(Down_tmp)):
        if Down< First_Capacity+Second_Capacity+Third_Capacity: #그 용량을 넘으면 안 됨.
            Down+=Down_tmp[i][0]
        else:
            break
    
    
    for i in range(len(First_All)):
        First_All[i][1]=First_All[i][1]*(1-in_advance)     #1-알파퍼센트 를 가지고 경쟁하려고 하는 준비
    
    for i in range(len(Second_All)):
        Second_All[i][1]=Second_All[i][1]*(1-in_advance)
    
    for i in range(len(Third_All)):
        Third_All[i][1]=Third_All[i][1]*(1-in_advance)    
        
     
    First_Sum = []
    for i in range(First_Number):
        if sum(First_Sum) <First_Capacity * 1.3-First_Middle_Sum:   #1.3배 한 후 미리 산 용량은 빼기
            try:
                First_Sum.append(First_All[i][1])
            except IndexError:
                break
        else:
            break
    First_Middle_Pass = First_All[:i+1]
    First_Middle_Fail = First_All[i+1:]
    
    for j in range(len(First_Middle_Pass)): #
        for k in range(j + 1, len(First_Middle_Pass)):
            if round((Upper_Limit-First_Middle_Pass[j][0])/Upper_Limit*70,2)+ First_Middle_Pass[j][2] < round((Upper_Limit-First_Middle_Pass[k][0])/Upper_Limit*70,2) + First_Middle_Pass[k][2]:
                First_Middle_Pass[j], First_Middle_Pass[k] = First_Middle_Pass[k], First_Middle_Pass[j]
    
    First_Sum = []
    for i in range(len(First_Middle_Pass) + 1):
        if sum(First_Sum) < First_Capacity-First_Middle_Sum:
            try:
                First_Sum.append(First_Middle_Pass[i][1])
            except IndexError:
                break
        else:
            break
    First_Final_Pass = First_Middle_Pass[:i+1]
    First_Final_Fail = First_Middle_Pass[i+1:]
    
    Second_All2 = Second_All + First_Middle_Fail + First_Final_Fail
    Second_All2.sort(reverse=True)
    
    Second_Sum = []
    for i in range(len(Second_All2)):
        if sum(Second_Sum) < Second_Capacity * 1.3-Second_Middle_Sum:
            try:
                Second_Sum.append(Second_All2[i][1])
            except IndexError:
                break
        else:
            break
    Second_Middle_Pass = Second_All2[:i]
    Second_Middle_Fail = Second_All2[i:]
    
    for j in range(len(Second_Middle_Pass)):
        for k in range(j + 1, len(Second_Middle_Pass)):
            if round((Upper_Limit-Second_Middle_Pass[j][0])/Upper_Limit*70,2)+ Second_Middle_Pass[j][2] < round((Upper_Limit-Second_Middle_Pass[k][0])/Upper_Limit*70,2) + Second_Middle_Pass[k][2]:
                Second_Middle_Pass[j], Second_Middle_Pass[k] = Second_Middle_Pass[k], Second_Middle_Pass[j]
    
    Second_Sum = []
    for i in range(len(Second_Middle_Pass) + 1):
        if sum(Second_Sum) < Second_Capacity-Second_Middle_Sum:
            try:
                Second_Sum.append(Second_Middle_Pass[i][1])
            except IndexError:
                break
        else:
            break
    Second_Final_Pass = Second_Middle_Pass[:i]
    Second_Final_Fail = Second_Middle_Pass[i:]
    
    Third_All2 = Third_All + Second_Middle_Fail + Second_Final_Fail
    Third_All2.sort(reverse=True)
    
    Third_Sum = []
    for i in range(len(Third_All2)):
        if sum(Third_Sum) < Third_Capacity * 1.3-Third_Middle_Sum:
            try:
                Third_Sum.append(Third_All2[i][1])
            except IndexError:
                break
        else:
            break
    Third_Middle_Pass = Third_All2[:i]
    Third_Middle_Fail = Third_All2[i:]
    
    for j in range(len(Third_Middle_Pass)):
        for k in range(j + 1, len(Third_Middle_Pass)):
            if round((Upper_Limit-Third_Middle_Pass[j][0])/Upper_Limit*70,2)+ Third_Middle_Pass[j][2] < round((Upper_Limit-Third_Middle_Pass[k][0])/Upper_Limit*70,2) + Third_Middle_Pass[k][2]:
                Third_Middle_Pass[j], Third_Middle_Pass[k] = Third_Middle_Pass[k], Third_Middle_Pass[j]
    
    Third_Sum = []
    for i in range(len(Third_Middle_Pass) + 1):
        if sum(Third_Sum) < Third_Capacity-Third_Middle_Sum:
            try:
                Third_Sum.append(Third_Middle_Pass[i][1])
            except IndexError:
                break
        else:
            break
    Third_Final_Pass = Third_Middle_Pass[:i]
    Third_Final_Fail = Third_Middle_Pass[i:]
    
    First_All_Fail = sorted(First_Middle_Fail[:] + First_Final_Fail[:], reverse=True)
    Second_All_Fail = sorted(Second_Middle_Fail[:] + Second_Final_Fail[:], reverse=True)
    Third_All_Fail = sorted(Third_Middle_Fail[:] + Third_Final_Fail[:], reverse=True)
    
    """print()
    print("First Final passes : ")
    print("정량적 가격           용량                정성적 평가")
    for i in range(len(First_Final_Pass)):
        print(First_Final_Pass[i])
        print()
    print()
    
    print("Second Final passes : ")
    print("정량적 가격           용량                정성적 평가")
    for i in range(len(Second_Final_Pass)):
        print(Second_Final_Pass[i])
        print()
    print()
    
    print("Third Final passes : ")
    print("정량적 가격           용량                정성적 평가")
    for i in range(len(Third_Final_Pass)):
        print(Third_Final_Pass[i])
        print()
    print()
    
    print("---------------------------------------------------------------------")"""
    
    Final_Pass = sorted(First_Final_Pass + Second_Final_Pass + Third_Final_Pass, reverse=True)
    First_Get = []
    Second_Get = []
    Third_Get = []
    
    for i in range(len(Final_Pass)):
        if Final_Pass[i] in First_All:
            First_Get.append(Final_Pass[i])
        elif Final_Pass[i] in Second_All:
            Second_Get.append(Final_Pass[i])
        else:
            Third_Get.append(Final_Pass[i])
    
    """print()
    print("First Get : ")
    print("정량적 가격           용량                정성적 평가")
    for i in range(len(First_Get)):
        print(First_Get[i])
        print()
    print()
    
    print("Second Get : ")
    print("정량적 가격           용량                정성적 평가")
    for i in range(len(Second_Get)):
        print(Second_Get[i])
        print()
    print()
    
    print("Third Get : ")
    print("정량적 가격           용량                정성적 평가")
    for i in range(len(Third_Get)):
        print(Third_Get[i])
        print()
    print()"""
    
    First_Xi=[]
    for i in range(len(First_Get)):
        First_Xi.append(1)
    for i in range(len(First_All)-len(First_Get)):
        First_Xi.append(in_advance)
    
    
    First_Up=sum(First_Xi)**2
    First_Down=len(First_Xi)*sum(i**2 for i in First_Xi)
    First_Fairness_Index=0 if First_Up==0 or First_Down==0 else First_Up/First_Down
    
    
    
    Second_Xi=[]
    for i in range(len(Second_Get)):
        Second_Xi.append(1)
    for i in range(len(Second_All)-len(Second_Get)):
        Second_Xi.append(in_advance)
    
    
    Second_Up=sum(Second_Xi)**2
    Second_Down=len(Second_Xi)*sum(i**2 for i in Second_Xi)
    Second_Fairness_Index=0 if Second_Up==0 or Second_Down==0 else Second_Up/Second_Down
    
    
    
    
    Third_Xi=[]
    for i in range(len(Third_Get)):
        Third_Xi.append(1)
    for i in range(len(Third_All)-len(Third_Get)):
        Third_Xi.append(in_advance)
    
    
    Third_Up=sum(Third_Xi)**2
    Third_Down=len(Third_Xi)*sum(i**2 for i in Third_Xi)
    Third_Fairness_Index=0 if Third_Up==0 or Third_Down==0 else Third_Up/Third_Down
    
    All_Fairness_Index=(First_Fairness_Index+Second_Fairness_Index+Third_Fairness_Index)/3
    """print(First_Fairness_Index)
    print(Second_Fairness_Index)
    print(Third_Fairness_Index)
    print(All_Fairness_Index)"""
    All_Average+=All_Fairness_Index
    First_Middle_Index+=First_Fairness_Index
    Second_Middle_Index+=Second_Fairness_Index
    Third_Middle_Index+=Third_Fairness_Index
    
    Up=0
    for i in range(len(First_Get)):
        Up+=First_Get[i][0]
    for i in range(len(Second_Get)):
        Up+=Second_Get[i][0]
    for i in range(len(Third_Get)):
        Up+=Third_Get[i][0]
    
    All_get=[]
    
    for i in range(len(First_Get)):
        All_get.append(First_Get[i][0])
    for i in range(len(Second_Get)):
        All_get.append(Second_Get[i][0])
    for i in range(len(Third_Get)):
        All_get.append(Third_Get[i][0])
    
    
    Up = Up +in_advance*(First_Middle_Sum+ Second_Middle_Sum+Third_Middle_Sum)    #이미 Up에다가 in_advance를 더함
    #print(p,First_Middle_Index,Second_Middle_Index,Third_Middle_Index)
    """ws.cell(row=p+2,column=o+1).value=First_Middle_Index
    ws.cell(row=p+15,column=o+1).value=Second_Middle_Index
    ws.cell(row=p+28,column=o+1).value=Third_Middle_Index
    ws.cell(row=p+41,column=o+1).value=(First_Middle_Index+Second_Middle_Index+Third_Middle_Index)/3"""
    #print(First_Middle_Index,Second_Middle_Index,Third_Middle_Index)
    #print(Up/Down) #효율성
    K.append(Up/Down)
    First_Middle_Index,Second_Middle_Index,Third_Middle_Index=0,0,0
wb.save("연습1.xlsx")
wb.close()
print(sum(K)/10)



