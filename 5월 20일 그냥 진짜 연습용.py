import openpyxl 

wb=openpyxl.load_workbook("연습.xlsx")
ws=wb.active
ws["B1"],ws["B14"]="가격 표준편차","가격 표준편차"
ws["B27"]="가격 표준편차"
ws["A2"]="용량 표준편차"
ws["A15"]="용량 표준편차"
ws["A28"]="용량 표준편차"
First_Middle_Index=123
Second_Middle_Index=123
Third_Middle_Index=123
for i in range(2,12):
    ws.cell(row=2,column=i).value=i-1
for i in range(2,12):
    ws.cell(row=15,column=i).value=i-1
for i in range(2,12):
    ws.cell(row=28,column=i).value=i-1

for j in range(3,13):
    ws.cell(row=j,column=1).value=j-2
for j in range(3,13):
    ws.cell(row=j+13,column=1).value=j-2
for j in range(3,13):
    ws.cell(row=j+26,column=1).value=j-2
for p in range(1,11):
    for o in range(1,11):
        ws.cell(row=j+26,column=1).value=j-2
        ws.cell(row=p+2,column=o+1).value=First_Middle_Index
        ws.cell(row=p+15,column=o+1).value=Second_Middle_Index
        ws.cell(row=p+28,column=o+1).value=Third_Middle_Index

wb.save("연습1.xlsx")
